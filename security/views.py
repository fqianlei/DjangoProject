import datetime
from datetime import date
import tushare as ts
from django.shortcuts import render
from django.http import HttpResponse
from django.db import connection
from django.template import loader
from security.models import BasicInfo, Forecast, MyFavorite,CCTVNews,AnnData
from django.core.paginator import Paginator
import openpyxl
import numpy as np
from django.contrib.auth.decorators import login_required
import os
from HTMLTable import HTMLTable
from django.utils.safestring import mark_safe   #用于转换HTML标签可以返回给浏览器直接显示
import time
import subprocess
import re
from django.core.exceptions import ObjectDoesNotExist
import pandas
#pip install html-table -i https://pypi.tuna.tsinghua.edu.cn/simple    安装HTMLTable
# Create your views here.  m = MyFavorite.objects.values('code')
#f = MyFavorite.objects.values_list('code',flat=True)
# fengql 123456;test  abc.1234
"""

class IndexView(generic.ListView):
    context_object_name = 'all_perforType_list'
    model = PerforType
    template_name = 'index.html'
"""
@login_required()
def index(request):
    with connection.cursor() as cursor:
        context_object_name = cursor.execute('''select f.perforType,count(id),notkanguo.newp from security_forecast f
                    left join (select ff.perforType,count(ff.id) newp from security_forecast ff where ff.kanguo is null group  by ff.perforType) notkanguo 
                    on notkanguo.perforType = f.perforType group by f.perforType''').fetchall()
    context = {
        'context_object_name': context_object_name,
    }
    return render(request, 'index.html', context)

@login_required()
def forcast(request, perforType):
    #all_forcast_list = Forecast.objects.filter(perforType__exact=perforType,kanguo=None,trade).order_by('-annDate')
    all_forcast_list = Forecast.objects.raw('''select * from security_forecast f 
    where f.kanguo is null and perforType=%s order by f.annDate desc,f.code''',[perforType])
    paginator = Paginator(all_forcast_list, 30)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'focastlist.html', {'page_obj': page_obj, 'perforType': perforType})

@login_required()
def addReason(request):
    if len(request.POST['myReason'])>2:
        code = BasicInfo.objects.get(pk=str(request.POST['code']))
        myf = MyFavorite(code=code, myReason=request.POST['myReason'])
        myf.save()
    if int(request.POST['fid'])!=99999:
        forcasta = Forecast.objects.get(pk = str(request.POST['fid']))
        forcasta.kanguo = '是'
        forcasta.save()
    return HttpResponse("<script>alert('保存成功');window.opener=null;window.top.open('','_self','');window.close(this);</script>")

@login_required()
def toAddReason(request):
    return render(request, 'searchresult.html')



@login_required()
def deleteMyfavorate(request):
    MyFavorite.objects.get(pk = request.POST['fid']).delete()
    return HttpResponse("<script>alert('删除成功');window.opener=null;window.top.open('','_self','');window.close(this);</script>")

def cctvnews(request):
    yesterday = str(date.today() - datetime.timedelta(days=1)).replace('-', '')
    with connection.cursor() as cursor:
        yesnews = cursor.execute('select * from security_cctvnews where date=%s',[yesterday]).fetchone()
        if yesnews:
            print("新闻已存在，无需重复添加")
        else:
            ts.set_token('619aac806adb72235ee6feb086f2cbb17cdeb4c85322e75c4f2f7e5d')
            pro = ts.pro_api()
            data = pro.cctv_news(date=yesterday)
            train_data = np.array(data)  # 先将数据框转换为数组
            train_data_list = train_data.tolist()  # 其次转换为列表
            #print(train_data_list)
            for i in range(len(train_data_list)):
                cctv =CCTVNews(date=train_data_list[i][0],title=train_data_list[i][1],content=train_data_list[i][2],)
                print('添加第',i,'条新闻')
                cctv.save()

    cctvnews = CCTVNews.objects.all().order_by('-id')
    paginator = Paginator(cctvnews, 8)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'cctvnews.html', {'page_obj': page_obj})

@login_required()
def insertForcast(request):
    wb = openpyxl.load_workbook('security/业绩预告.xlsx')
    sheet = wb.worksheets[0]
    rows = sheet.max_row
    daoruhang = 0
    for i in range(2, rows):
        annDate = sheet['A' + str(i)].value
        code = sheet['B' + str(i)].value
        annPeriod = sheet['E' + str(i)].value
        perforType = sheet['F' + str(i)].value
        name = sheet['C' + str(i)].value
        with connection.cursor() as cursor:
            context_object_name = cursor.execute('select * from security_forecast where annDate=%s and code=%s and annPeriod=%s and perforType=%s',[annDate,code,annPeriod,perforType]).fetchone()
            if context_object_name:
                pass
            else:
                searchname = cursor.execute('select name from security_basicinfo where code=%s',[code]).fetchone()
                if searchname:
                    # cursor.execute('insert into security_basicinfo (code,name) VALUES (%s,%s)', [code,'test'])
                    trade = sheet['D' + str(i)].value
                    perforContent = sheet['G' + str(i)].value
                    changeReason = sheet['H' + str(i)].value
                    upperLimit = sheet['I' + str(i)].value
                    lowerLimit = sheet['J' + str(i)].value
                    sheetForcast = [annDate, code, name, trade, annPeriod, perforType, perforContent, changeReason,
                                    upperLimit, lowerLimit, datetime.datetime.now()]
                    daoruhang = daoruhang + 1
                    with connection.cursor() as cursor:
                        cursor.execute('insert into security_forecast (annDate,code,name,trade,annPeriod,perforType,perforContent,changeReason,upperLimit,lowerLimit,addtime) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',sheetForcast)
                        #cursor.execute('delete  from security_forecast where  code not in (select code from security_basicinfo)')
                else:
                    pass
    print('本次共导入',daoruhang , '条业绩预告')
    return HttpResponse("<script>alert('更新业绩预告成功');window.opener=null;window.top.open('','_self','');window.close(this);</script>")

#两个EXCEL，先导入业绩预告 ，再导入ROE等数据
@login_required()
def insertRoe(request):
    start =time.perf_counter()
    #os.system("python security/stockbasic.py")   通过toshare获取股票信息
    # 导入ROE等内容，先删除股票名称，再全部复制粘贴EXCEL，主要是去除双引号
    wb = openpyxl.load_workbook('security/股票基本信息.xlsx')
    sheet = wb.worksheets[0]
    rows = sheet.max_row
    stocklist = []
    for i in range(2, rows):
        code = sheet['B' + str(i)].value
        name = sheet['C' + str(i)].value
        PE = sheet['D' + str(i)].value # 动态市盈率
        PB = sheet['E' + str(i)].value # 市净率
        GMV = sheet['F' + str(i)].value # 总市值
        province = sheet['G' + str(i)].value  # 省份
        city = sheet['H' + str(i)].value  # 城市
        webSite = sheet['I' + str(i)].value  # 网站
        industry = sheet['J' + str(i)].value  # 行业
        business_scope = sheet['K' + str(i)].value  # 经营范围
        market = sheet['l' + str(i)].value  # 上市地点
        list_date = sheet['M' + str(i)].value  # 上市地点
        #以下指标放入主表，主要是开发方便，使用灵活，可以根据时间不同查看不同时间的财务数据
        BI = sheet['N' + str(i)].value  # Business Income 营业收入
        OP = sheet['O' + str(i)].value  # operating profit 营业利润
        GPM = sheet['P' + str(i)].value  # 销售毛利率 Gross profit margin
        NPRA = sheet['Q' + str(i)].value  # 销售净利率Net profit rate of sales
        YOY = sheet['R' + str(i)].value  # 营业收入（同比增长率）
        AR = sheet['S' + str(i)].value  # #应收账款Accounts receivable
        #onestock = [code,name,PE,PB,GMV,province,city,webSite,industry,business_scope,market]
        try:
            basicinfo = BasicInfo.objects.get(code=code)
            basicinfo.name = name
            basicinfo.PE = PE
            basicinfo.PB = PB
            basicinfo.GMV = GMV
            basicinfo.province = province
            basicinfo.city = city
            basicinfo.webSite = webSite
            basicinfo.industry = industry
            basicinfo.business_scope = business_scope
            basicinfo.market = market
            basicinfo.list_date = list_date
            basicinfo.BI = BI
            basicinfo.OP = OP
            basicinfo.GPM = GPM
            basicinfo.NPRA = NPRA
            basicinfo.YOY = YOY
            basicinfo.AR = AR
            basicinfo.save()
        except ObjectDoesNotExist:
            BasicInfo(code=code,name=name,PE=PE,PB=PB,GMV=GMV,
                      province=province,city=city,webSite=webSite,industry=industry,
                      business_scope=business_scope,list_date=list_date,
                      BI=BI,OP=OP,GPM=GPM,NPRA=NPRA,YOY=YOY,AR=AR,).save()
    end = time.perf_counter()
    print('执行时间: %s 秒' % (end - start))
    #BasicInfo.objects.all().delete()
    #BasicInfo.objects.bulk_create(stocklist)
    return HttpResponse("<script>alert('更新股票基本信息和ROE等动态数据成功');window.opener=null;window.top.open('','_self','');window.close(this);</script>")


@login_required()
def readByline(request):
    all_mycode_list = BasicInfo.objects.raw('select * from security_basicinfo where code in (select code from security_myfavorite)')
    paginator = Paginator(all_mycode_list, 15)
    #| date:"Y-m-d H:i:s"
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'readByline.html', {'page_obj': page_obj,})

@login_required()
def insertAnndata(request):
    start =time.perf_counter()
    wb = openpyxl.load_workbook('security/anndata.xlsx')
    sheet = wb.worksheets[0]
    rows = sheet.max_row
    column =sheet.max_column
    #开始按列循环更新财务数据
    for i in range(4, column+1):
        annType = sheet.cell(row=1,column=i).value.split('-')[0]
        annPeriod = sheet.cell(row=1,column=i).value.split('-')[1]
        for j in range(2, rows + 1):
            code = sheet['B' + str(j)].value
            annResult = sheet.cell(row=j,column=i).value
            try:
                tempData = AnnData.objects.get(annType=annType, annPeriod=annPeriod, code=code)
                if tempData.annResult == '-':
                    tempData.annResult = annResult
                    tempData.save()
            except ObjectDoesNotExist:
                if BasicInfo.objects.filter(code=code).count() > 0:
                    AnnData(code=BasicInfo.objects.get(pk=code),annType=annType,annPeriod=annPeriod,annResult=annResult).save()
    end = time.perf_counter()
    print('执行时间: %s 秒' % (end - start))
    return HttpResponse("<script>alert('更新财务数据成功');window.opener=null;window.top.open('','_self','');window.close(this);</script>")

@login_required()
def tosearch(request):
    return render(request, 'tosearch.html')

@login_required()
def tosearch2(request):
    return render(request, 'tosearch2.html')

@login_required()
def tosearch3(request):
    return render(request, 'tosearch3.html')

@login_required()
def search(request):
    listCode = turntocodelist(request.POST['code'])
    basicinfo = BasicInfo.objects.filter(code__in=listCode).order_by('GMV')
    return render(request, 'searchresult.html', {'basicinfo': basicinfo})

@login_required()
def search2(request):
    #zhibiao 财务指标
    codelist = turntocodelist(request.POST['code'])
    zhibiao = ''
    for i in range(0,len(codelist)):
        temp = niubi(codelist[i])
        zhibiao = zhibiao+temp
    zhibiao = mark_safe(zhibiao)  #转换后才能显示正确
    context = {
        'zhibiao':zhibiao,
    }
    template = loader.get_template('searchresult2.html')
    return HttpResponse(template.render(context,request))

def search3(request):
    #zhibiao 财务指标
    codelist = turntocodelist(request.POST['code'])
    zhibiao = mark_safe(niubi2(codelist))
    context = {
        'zhibiao':zhibiao,
    }
    template = loader.get_template('searchresult2.html')
    return HttpResponse(template.render(context,request))

#用于单个查询用：包含多个财务周期I数据
def zhibiao(request,code):
    zhibiao = niubi(code)
    zhibiao = mark_safe(zhibiao)  #转换后才能显示正确
    context = {
        'zhibiao':zhibiao,
    }
    template = loader.get_template('searchresult2.html')
    return HttpResponse(template.render(context,request))

#用于单个查询用，包含评分数据
def zhibiao2(request,code):
    codelist = turntocodelist(code)
    zhibiao = mark_safe(niubi2(codelist))
    context = {
        'zhibiao':zhibiao,
    }
    template = loader.get_template('searchresult2.html')
    return HttpResponse(template.render(context,request))


def niubi(code):
    #适用于单个公司和多个公司的多个周期财务报表显示
    annPeriod = ['20171231','20181231','20191231','20200331','20200630','20200930']
    annType = ['营业收入','营收增幅','营业利润','净利润','应收账款','销售毛利率','销售净利率','ROE','资产负债率','总资产周转率','预付账款','总资产增长率','流动比率','销售期间费用率','管理费用率','营业费用率','财务费用率','非流动资产/总资产']
    # '扣非净利润增幅'指标有问题，所有没有显示
    basic = BasicInfo.objects.get(pk=code)
    table = HTMLTable(caption=str(code) +'：'+ basic.name+'-市盈率：'+basic.PE+'-市值：'+basic.GMV)
    table.append_header_rows((
        ['序号','指标','20171231','20181231','20191231','20200331','20200630','20200930','指标'],
    ))
    for i in range(0,len(annType)):
        addrow = (annType[i],
                   )
        for j in range(0, len(annPeriod)):
            try:
                tempData = AnnData.objects.get(annType=annType[i], annPeriod=annPeriod[j], code=code)
                temp = (tempData.annResult,
                    )
            except ObjectDoesNotExist:
                temp = (
                    '-',
                )
            addrow = addrow + temp
        addrow = (i+1,)+addrow + (annType[i],) #右侧也有指标，这样看起来方便；(i+1,)是为了显示序号； (annType[i],)是未来在右侧显示指示表名称，方便查看
        table.append_data_rows((
         addrow,
        ))
    #下面两个结合才可以显示正确的表格样式
    table.set_cell_style({
        'border-style': 'solid',
        'padding': '1px',
        'spacing': '0px',
    })
    table.set_style({
        'border-collapse': 'collapse',
        'word-break': 'keep-all',
        'white-space': 'nowrap',
        'font-size': '14px',
    })
    html = table.to_html()
    return html

def niubi2(codelist):
    #适用于多个公司的单个周期
    annPeriod = '20200930'
    annType = ['营业收入','营收增幅','营业利润','净利润','应收账款','销售毛利率','销售净利率','ROE','资产负债率','总资产周转率','预付账款','总资产增长率','流动比率','销售期间费用率','管理费用率','营业费用率','财务费用率','非流动资产/总资产']
    # '扣非净利润增幅'指标有问题，所有没有显示
    header_rows = ['序号','指标']
    print('++++++++++++++++++++++++++++++++++++++++++++++++++++')
    print(codelist)
    for i in range(0,len(codelist)):
        company = codelist[i]
        basic = BasicInfo.objects.get(pk=codelist[i])
        cname = basic.name
        title = cname+':'+str(company)
        header_rows.append(title)
    header_rows.append('指标')

    table = HTMLTable('财务年度'+annPeriod)
    table.append_header_rows((
        header_rows,
    ))

    #增加市值PE
    addrow = [
        1,
        'PE',
    ]
    for i in range(0,len(codelist)):
        basic = BasicInfo.objects.get(pk=codelist[i])
        addrow.append(basic.PE)
    addrow.append('PE')
    table.append_data_rows((
        addrow,
    ))
    #增加市值GMV
    addrow = [
        2,
        'GMV总市值',
    ]
    for i in range(0,len(codelist)):
        basic = BasicInfo.objects.get(pk=codelist[i])
        addrow.append(basic.GMV)
    addrow.append('GMV总市值')
    table.append_data_rows((
        addrow,
    ))

    for i in range(0,len(annType)):
        addrow = (annType[i],
                  )
        for j in range(0,len(codelist)):
            try:
                tempData = AnnData.objects.get(annType=annType[i], annPeriod=annPeriod, code=codelist[j])
                temp = (tempData.annResult,
                    )
            except ObjectDoesNotExist:
                temp = (
                    '-',
                )
            addrow = addrow + temp
        addrow = (i + 3,) + addrow + (annType[i],)  # 右侧也有指标，这样看起来方便；(i+1,)是为了显示序号； (annType[i],)是未来在右侧显示指示表名称，方便查看
        table.append_data_rows((
            addrow,
        ))

    #评分模型
    addrow = [
        len(annType)+3,
        '综合得分',
    ]
    for i in range(0,len(codelist)):
        basic = BasicInfo.objects.get(pk=codelist[i])
        ROE = AnnData.objects.get(annType='ROE', annPeriod=annPeriod, code=codelist[i]).annResult
        ROE =float(ROE)
        print('ROE:',ROE)
        if ROE>=15:
            roeScore = 2
        elif ROE>=10 and ROE<15:
            roeScore = 1
        elif ROE >= 5 and ROE < 10:
            roeScore = 0
        else:
            roeScore = -1
        netProfitMargin = AnnData.objects.get(annType='销售净利率', annPeriod=annPeriod, code=codelist[i]).annResult
        netProfitMargin = float(netProfitMargin)
        print('netProfitMargin:', netProfitMargin)
        if netProfitMargin>=15:
            netProfitMarginScore = 2
        elif netProfitMargin>=10 and netProfitMargin<15:
            netProfitMarginScore = 1
        elif netProfitMargin >= 5 and netProfitMargin < 10:
            netProfitMarginScore = 0
        else:
            netProfitMarginScore = -1
        AR = AnnData.objects.get(annType='应收账款', annPeriod=annPeriod, code=codelist[i]).annResult
        BI = AnnData.objects.get(annType='营业收入', annPeriod=annPeriod, code=codelist[i]).annResult
        ab = float(AR)/float(BI)
        print('ab:', ab)
        if ab>=0.3:
            abScore = -1
        elif ab>=0.2 and ab<0.3:
            abScore = 0
        elif ab >= 0.1 and ab < 0.2:
            abScore = 1
        else:
            abScore = 2
        YOY = AnnData.objects.get(annType='营收增幅', annPeriod=annPeriod, code=codelist[i]).annResult
        YOY = float(YOY)
        print('YOY:', YOY)
        if YOY>=20:
            YOYScore = 2
        elif YOY>=10 and YOY<20:
            YOYScore = 1
        elif YOY >= 5 and YOY < 10:
            YOYScore = 0
        else:
            YOYScore = -1
        basic = BasicInfo.objects.get(pk=codelist[i])
        YOY = AnnData.objects.get(annType='营收增幅', annPeriod=annPeriod, code=codelist[i]).annResult
        yb =float(YOY)/float(basic.PE)
        print('yb:', yb)
        if yb>=2:
            ybScore = 2
        elif yb>=1 and yb<2:
            ybScore = 1
        elif yb >= 1 and yb < 1:
            ybScore = 0
        else:
            ybScore = -1
        totalScore = roeScore + netProfitMarginScore + abScore + YOYScore + ybScore
        print('totalScore:', totalScore)
        addrow.append(str(totalScore))
    addrow.append('综合得分')
    table.append_data_rows((
        addrow,
    ))

    #下面两个结合才可以显示正确的表格样式
    table.set_cell_style({
        'border-style': 'solid',
        'padding': '1px',
        'spacing': '0px',
    })
    table.set_style({
        'border-collapse': 'collapse',
        'word-break': 'keep-all',
        'white-space': 'nowrap',
        'font-size': '14px',
    })
    html = table.to_html()
    return html



#获取查询结果并转换为CODELIST的公共函数
def turntocodelist(codelist):
    #将字符串根据换行符切割为列表
    listCode1 = codelist.split('\r\n')
    listCode = []
    for i in range(len(listCode1)):
        if len(listCode1[i])==6:
            listCode.append(listCode1[i])
            #取值包含证券市场的代码，如：000411.SZ
        elif len(listCode1[i])==9:
            listCode.append(listCode1[i][0:6])
            print(listCode1[i][0:6])
        else:
            if len(listCode1[i])>2:
                if listCode.append(BasicInfo.objects.get(name = listCode1[i])):
                    listCode.append(BasicInfo.objects.get(name=listCode1[i]).code)
    return listCode


def printipv6(request):
    child = subprocess.Popen("ifconfig", shell=True, stdout=subprocess.PIPE)
    out = child.communicate();  # 保存ipconfig中的所有信息
    ipv6_pattern = '(([a-f0-9]{1,4}:){7}[a-f0-9]{1,4})'
    m = re.findall(ipv6_pattern, str(out))
    address = m[1][0]
    context = {
        'address':address,
    }
    template = loader.get_template('ipv6.html')
    return HttpResponse(template.render(context,request))




